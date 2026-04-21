#!/usr/bin/env python3
"""
app.py — Flask application and all API routes
عيادة النور البصرية — Optical Clinic Management System

Database : Supabase (PostgreSQL)
Deployment: Vercel (WSGI)  |  local: python start.py
"""

import os
import io
import json
import zipfile
import traceback
import urllib.request
import urllib.error
from datetime import datetime, date, timedelta

from flask import Flask, request, jsonify, send_from_directory, send_file
from supabase import create_client, Client

# ── App setup ──────────────────────────────────────────────────────────────────
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR = os.path.join(BASE_DIR, "static")
os.makedirs(STATIC_DIR, exist_ok=True)

app = Flask(__name__, static_folder=STATIC_DIR)
app.config["MAX_CONTENT_LENGTH"] = 32 * 1024 * 1024   # 32 MB upload limit

# ── Supabase client ────────────────────────────────────────────────────────────
SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://ekzvzaviejiovwnixpcq.supabase.co")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "sb_publishable_DcUzLlY-_wCeD78pZTGOSg_ByR82M5d")

_supabase: Client = None

def get_db() -> Client:
    global _supabase
    if _supabase is None:
        _supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
    return _supabase

def init_db():
    pass   # tables live in Supabase — run schema.sql once in the SQL editor

# ── Type-safe coercion helpers ─────────────────────────────────────────────────

def _str(v, default=""):
    if v is None: return default
    return str(v).strip()

def _int(v, default=None):
    try:    return int(v) if v not in (None, "", "null") else default
    except: return default

def _float(v, default=0.0):
    try:    return float(v) if v not in (None, "", "null") else default
    except: return default

def _bool(v, default=False):
    if isinstance(v, bool): return v
    if isinstance(v, int):  return bool(v)
    if isinstance(v, str):  return v.strip().lower() in ("1","true","yes")
    return default

def today():
    return date.today().isoformat()

def calc_next_review(d=None):
    try:
        dt = datetime.strptime(d or today(), "%Y-%m-%d")
        m  = dt.month + 6
        y  = dt.year + (m - 1) // 12
        m  = (m - 1) % 12 + 1
        dt = dt.replace(year=y, month=m)
    except Exception:
        dt = datetime.now() + timedelta(days=180)
    return dt.strftime("%Y-%m-%d")

def sb_rows(response):
    return response.data or []

def sb_one(response):
    data = response.data
    return data[0] if data else None

def err(e):
    tb = traceback.format_exc()
    print("ERROR:", tb)
    return jsonify({"error": str(e), "detail": tb}), 400


# ══ STATIC ════════════════════════════════════════════════════════════════════

@app.route("/")
def index():
    return send_from_directory(STATIC_DIR, "index.html")


# ══ PATIENTS ══════════════════════════════════════════════════════════════════

@app.route("/api/patients", methods=["GET"])
def get_patients():
    try:
        rows = sb_rows(
            get_db().table("patients").select("*")
              .order("admission_date", desc=True).order("id", desc=True).execute()
        )
        return jsonify(rows)
    except Exception as e:
        return err(e)


@app.route("/api/patients", methods=["POST"])
def save_patient():
    try:
        d   = request.json or {}
        pid = _int(d.get("id"))
        db  = get_db()

        vals = {
            "name"          : _str(d.get("name")),
            "age"           : _str(d.get("age")),
            "phone"         : _str(d.get("phone")),
            "address"       : _str(d.get("address")),
            "admission_date": _str(d.get("admission_date")) or today(),
            "next_review"   : calc_next_review(_str(d.get("admission_date")) or None),
            "notes"         : _str(d.get("notes")),
            "od_sph": _str(d.get("od_sph")),  "od_cyl": _str(d.get("od_cyl")),
            "od_axis": _str(d.get("od_axis")), "od_add": _str(d.get("od_add")),
            "od_va":  _str(d.get("od_va")),    "od_bcva": _str(d.get("od_bcva")),
            "os_sph": _str(d.get("os_sph")),   "os_cyl": _str(d.get("os_cyl")),
            "os_axis": _str(d.get("os_axis")), "os_add": _str(d.get("os_add")),
            "os_va":  _str(d.get("os_va")),    "os_bcva": _str(d.get("os_bcva")),
            "pd"           : _str(d.get("pd")),
            "checkup_done" : 1 if _bool(d.get("checkup_done"))  else 0,
            "checkup_paid" : 1 if _bool(d.get("checkup_paid"))  else 0,
            "checkup_fee"  : _float(d.get("checkup_fee"),  5000.0),
            "lenses_qty"   : _int(d.get("lenses_qty"),     2),
            "lens1_id"     : _int(d.get("lens1_id")),
            "lens1_cost"   : _float(d.get("lens1_cost"),   0.0),
            "lens2_id"     : _int(d.get("lens2_id")),
            "lens2_cost"   : _float(d.get("lens2_cost"),   0.0),
            "frame_price"  : _float(d.get("frame_price"),  0.0),
            "frame_cost"   : _float(d.get("frame_cost"),   0.0),
            "total_cost"   : _float(d.get("total_cost"),   0.0),
            "amount_paid"  : _float(d.get("amount_paid"),  0.0),
            "frame_name"   : _str(d.get("frame_name")),
            "frame_type"   : _str(d.get("frame_type")),
            "rxtype"       : _str(d.get("rxtype")),
            "material"     : _str(d.get("material")),
            "features"     : _str(d.get("features")),
        }

        new_l1 = vals["lens1_id"]
        new_l2 = vals["lens2_id"]
        lqty   = vals["lenses_qty"] or 2

        if pid:
            old = sb_one(db.table("patients").select("lens1_id,lens2_id,lenses_qty")
                           .eq("id", pid).execute())
            old_l1 = _int(old.get("lens1_id")) if old else None
            old_l2 = _int(old.get("lens2_id")) if old else None
            db.table("patients").update(vals).eq("id", pid).execute()
            if old_l1 and old_l1 != new_l1: _lens_adjust(db, old_l1, +1)
            if old_l2 and old_l2 != new_l2: _lens_adjust(db, old_l2, +1)
            if new_l1 and new_l1 != old_l1: _lens_adjust(db, new_l1, -1)
            if new_l2 and lqty == 2 and new_l2 != old_l2: _lens_adjust(db, new_l2, -1)
        else:
            res = db.table("patients").insert(vals).execute()
            row = sb_one(res)
            if not row:
                return jsonify({"error": "Insert returned no data — check Supabase RLS policies"}), 400
            pid = row["id"]
            if new_l1: _lens_adjust(db, new_l1, -1)
            if new_l2 and lqty == 2: _lens_adjust(db, new_l2, -1)
            if vals["total_cost"] > 0:
                db.table("ledger").insert({
                    "date": vals["admission_date"], "entry_type": "income",
                    "category": "مبيعات مرضى",
                    "description": f"بيع للمريض: {vals['name']}",
                    "total_amount": vals["total_cost"], "paid_amount": vals["amount_paid"],
                    "is_expense": False, "source_ref": f"patient:{pid}",
                }).execute()

        return jsonify({"id": pid})
    except Exception as e:
        return err(e)


def _lens_adjust(db, lid, delta):
    try:
        row = sb_one(db.table("lenses").select("stock").eq("id", lid).execute())
        if row:
            db.table("lenses").update(
                {"stock": max(0, (_int(row.get("stock"), 0) or 0) + delta)}
            ).eq("id", lid).execute()
    except Exception as e:
        print(f"_lens_adjust lid={lid} delta={delta}: {e}")


@app.route("/api/patients/<int:pid>", methods=["DELETE"])
def delete_patient(pid):
    try:
        db  = get_db()
        row = sb_one(db.table("patients").select("lens1_id,lens2_id,lenses_qty")
                       .eq("id", pid).execute())
        if row:
            if row.get("lens1_id"): _lens_adjust(db, row["lens1_id"], +1)
            if row.get("lens2_id") and (_int(row.get("lenses_qty"), 2) or 2) == 2:
                _lens_adjust(db, row["lens2_id"], +1)
        db.table("patients").delete().eq("id", pid).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return err(e)


@app.route("/api/patients/<int:pid>/pay", methods=["POST"])
def pay_patient(pid):
    try:
        d      = request.json or {}
        amount = _float(d.get("amount"))
        db     = get_db()
        row    = sb_one(db.table("patients").select("amount_paid").eq("id", pid).execute())
        if row:
            db.table("patients").update(
                {"amount_paid": _float(row.get("amount_paid")) + amount}
            ).eq("id", pid).execute()
        db.table("patient_payments").insert({
            "patient_id": pid, "amount": amount,
            "date": _str(d.get("date")) or today(),
        }).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return err(e)


# ══ DEBTORS ═══════════════════════════════════════════════════════════════════

@app.route("/api/debtors", methods=["GET"])
def get_debtors():
    try:
        return jsonify(sb_rows(get_db().table("debtors").select("*")
                                .order("created_at", desc=True).execute()))
    except Exception as e:
        return err(e)


@app.route("/api/debtors", methods=["POST"])
def save_debtor():
    try:
        d         = request.json or {}
        did       = _int(d.get("id"))
        total     = _float(d.get("total_we_owe"))
        init_paid = _float(d.get("initial_paid"))
        remaining = max(0.0, total - init_paid)
        status    = "settled" if remaining <= 0 else ("partial" if init_paid > 0 else "unpaid")
        db        = get_db()
        payload   = {
            "name": _str(d.get("name")), "phone": _str(d.get("phone")),
            "category": _str(d.get("category")), "what_bought": _str(d.get("what_bought")),
            "total_we_owe": total, "remaining": remaining, "status": status,
            "due_date": _str(d.get("due_date")) or None, "notes": _str(d.get("notes")),
        }
        if did:
            db.table("debtors").update(payload).eq("id", did).execute()
        else:
            payload["total_paid"] = init_paid
            res = db.table("debtors").insert(payload).execute()
            did = sb_one(res)["id"]
        return jsonify({"id": did})
    except Exception as e:
        return err(e)


@app.route("/api/debtors/<int:did>", methods=["DELETE"])
def delete_debtor(did):
    try:
        get_db().table("debtors").delete().eq("id", did).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return err(e)


@app.route("/api/debtors/<int:did>/pay", methods=["POST"])
def pay_debtor(did):
    try:
        d      = request.json or {}
        amount = _float(d.get("amount"))
        db     = get_db()
        row    = sb_one(db.table("debtors").select("total_paid,remaining").eq("id", did).execute())
        if row:
            new_paid      = _float(row.get("total_paid")) + amount
            new_remaining = max(0.0, _float(row.get("remaining")) - amount)
            db.table("debtors").update({
                "total_paid": new_paid, "remaining": new_remaining,
                "status": "settled" if new_remaining <= 0 else ("partial" if new_paid > 0 else "unpaid"),
            }).eq("id", did).execute()
        db.table("debtor_payments").insert({
            "debtor_id": did, "amount": amount,
            "method": _str(d.get("method"), "cash"),
            "note": _str(d.get("note")),
            "date": _str(d.get("date")) or today(),
        }).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return err(e)


# ══ LENSES ════════════════════════════════════════════════════════════════════

@app.route("/api/lenses", methods=["GET"])
def get_lenses():
    try:
        return jsonify(sb_rows(get_db().table("lenses").select("*")
                                .order("lens_type").order("sph").execute()))
    except Exception as e:
        return err(e)


@app.route("/api/lenses", methods=["POST"])
def save_lens():
    try:
        d = request.json or {}
        lid = _int(d.get("id"))
        db  = get_db()
        payload = {
            "lens_type": _str(d.get("lens_type")), "category": _str(d.get("category")),
            "sph": _str(d.get("sph")), "cyl": _str(d.get("cyl")),
            "material": _str(d.get("material")), "filter_type": _str(d.get("filter_type")),
            "tint": _str(d.get("tint")),
            "stock": _int(d.get("stock"), 0), "reorder_point": _int(d.get("reorder_point"), 5),
            "cost": _float(d.get("cost")), "price": _float(d.get("price")),
        }
        if lid: db.table("lenses").update(payload).eq("id", lid).execute()
        else:   db.table("lenses").insert(payload).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return err(e)


@app.route("/api/lenses/<int:lid>/restock", methods=["POST"])
def restock_lens(lid):
    try:
        d    = request.json or {}
        qty  = _int(d.get("qty"), 0)
        cost = _float(d.get("cost"))
        db   = get_db()
        row  = sb_one(db.table("lenses").select("*").eq("id", lid).execute())
        if not row: return jsonify({"error": "Lens not found"}), 404
        new_stock = (_int(row.get("stock"), 0) or 0) + qty
        upd = {"stock": new_stock}
        if cost > 0: upd["cost"] = cost
        db.table("lenses").update(upd).eq("id", lid).execute()
        db.table("lens_restocks").insert({
            "lens_id": lid, "qty": qty, "cost_per_unit": cost,
            "date": _str(d.get("date")) or today(),
        }).execute()
        if qty * cost > 0:
            db.table("ledger").insert({
                "date": _str(d.get("date")) or today(), "entry_type": "expense",
                "category": "شراء عدسات",
                "description": f"تجديد: {row.get('lens_type','')} SPH{row.get('sph','')} × {qty}",
                "total_amount": qty*cost, "paid_amount": qty*cost,
                "is_expense": True, "source_ref": f"restock:lens:{lid}",
            }).execute()
        return jsonify({"ok": True, "new_stock": new_stock})
    except Exception as e:
        return err(e)


@app.route("/api/lenses/<int:lid>", methods=["DELETE"])
def delete_lens(lid):
    try:
        get_db().table("lenses").delete().eq("id", lid).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return err(e)


# ══ OP-COSTS ══════════════════════════════════════════════════════════════════

@app.route("/api/op-costs", methods=["GET"])
def get_op_costs():
    try:
        return jsonify(sb_rows(get_db().table("op_costs").select("*")
                                .order("created_at", desc=True).execute()))
    except Exception as e:
        return err(e)


@app.route("/api/op-costs", methods=["POST"])
def save_op_cost():
    try:
        d = request.json or {}
        oid = _int(d.get("id"))
        db  = get_db()
        payload = {
            "name": _str(d.get("name")), "amount": _float(d.get("amount")),
            "frequency": _str(d.get("frequency")), "category": _str(d.get("category")),
        }
        if oid: db.table("op_costs").update(payload).eq("id", oid).execute()
        else:   db.table("op_costs").insert(payload).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return err(e)


@app.route("/api/op-costs/<int:oid>", methods=["DELETE"])
def delete_op_cost(oid):
    try:
        get_db().table("op_costs").delete().eq("id", oid).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return err(e)


# ══ SETTINGS ══════════════════════════════════════════════════════════════════

@app.route("/api/settings", methods=["GET"])
def get_settings():
    try:
        rows = sb_rows(get_db().table("settings").select("key,value").execute())
        return jsonify({r["key"]: r["value"] for r in rows})
    except Exception as e:
        return err(e)


@app.route("/api/settings", methods=["POST"])
def save_settings():
    try:
        d  = request.json or {}
        db = get_db()
        for k, v in d.items():
            db.table("settings").upsert(
                {"key": k, "value": str(v) if v is not None else ""},
                on_conflict="key"
            ).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return err(e)


# ══ DASHBOARD ═════════════════════════════════════════════════════════════════

@app.route("/api/dashboard", methods=["GET"])
def dashboard():
    try:
        db       = get_db()
        t        = today()
        week_ago = (date.today() - timedelta(days=7)).isoformat()
        today_pts = sb_rows(db.table("patients").select("total_cost,amount_paid")
                              .eq("admission_date", t).execute())
        week_pts  = sb_rows(db.table("patients").select("total_cost")
                              .gte("admission_date", week_ago).execute())
        all_pts   = sb_rows(db.table("patients").select("total_cost,amount_paid").execute())
        op_costs  = sb_rows(db.table("op_costs").select("amount").execute())
        we_owe    = sb_rows(db.table("debtors").select("remaining").gt("remaining", 0).execute())
        low_rows  = sb_rows(db.table("lenses").select("stock,reorder_point").execute())
        week_sales = sum(_float(p.get("total_cost")) for p in week_pts)
        total_oc   = sum(_float(r.get("amount"))     for r in op_costs)
        return jsonify({
            "today": {
                "sales":     sum(_float(p.get("total_cost"))  for p in today_pts),
                "collected": sum(_float(p.get("amount_paid")) for p in today_pts),
                "patients":  len(today_pts),
            },
            "week":     {"sales": week_sales, "net": week_sales - total_oc},
            "pt_debt":  sum(max(0.0, _float(p.get("total_cost")) - _float(p.get("amount_paid"))) for p in all_pts),
            "we_owe":   sum(_float(r.get("remaining")) for r in we_owe),
            "low_stock": sum(1 for r in low_rows if (_int(r.get("stock"),0) or 0) <= (_int(r.get("reorder_point"),5) or 5)),
        })
    except Exception as e:
        return err(e)


# ══ REPORTS ═══════════════════════════════════════════════════════════════════

@app.route("/api/reports", methods=["GET"])
def reports():
    try:
        start = request.args.get("start", (date.today() - timedelta(days=7)).isoformat())
        end   = request.args.get("end",   today())
        db    = get_db()

        # Patients in range
        pts = sb_rows(db.table("patients").select("*")
                        .gte("admission_date", start).lte("admission_date", end)
                        .order("admission_date", desc=True).execute())

        # Restocks in range
        restocks = sb_rows(db.table("lens_restocks").select("*")
                             .gte("date", start).lte("date", end)
                             .order("date", desc=True).execute())

        # Custom sales in range
        custom_sales = sb_rows(db.table("custom_sales").select("*")
                                 .gte("date", start).lte("date", end)
                                 .order("date", desc=True).execute())

        # Op costs
        op_costs_rows = sb_rows(db.table("op_costs").select("*").execute())

        try:
            days = max(1, (datetime.strptime(end,"%Y-%m-%d") - datetime.strptime(start,"%Y-%m-%d")).days + 1)
        except Exception:
            days = 7

        op_est = 0.0
        for r in op_costs_rows:
            amt  = _float(r.get("amount"))
            freq = _str(r.get("frequency"))
            if freq=="weekly":    op_est += amt*(days/7)
            elif freq=="monthly": op_est += amt*(days/30)
            elif freq=="yearly":  op_est += amt*(days/365)
            elif freq=="daily":   op_est += amt*days
            elif freq=="once":    op_est += amt/52

        # Revenue breakdown
        lens_rev     = sum((_float(p.get("lens1_cost")) + _float(p.get("lens2_cost"))) for p in pts)
        frame_rev    = sum(_float(p.get("frame_price") or p.get("frame_cost")) for p in pts)
        checkup_rev  = sum(_float(p.get("checkup_fee") or 0) for p in pts)
        custom_rev   = sum(_float(s.get("total")) for s in custom_sales)
        gross_rev    = lens_rev + frame_rev + checkup_rev + custom_rev
        collected    = sum(_float(p.get("amount_paid")) for p in pts)
        restock_cost = sum((_float(r.get("qty",0)) * _float(r.get("cost_per_unit",0))) for r in restocks)
        gross_profit = round(gross_rev - restock_cost - op_est, 2)

        return jsonify({
            "patients":     pts,
            "restocks":     restocks,
            "custom_sales": custom_sales,
            "op_costs":     op_costs_rows,
            "stats": {
                "gross_revenue":   round(gross_rev, 2),
                "collected":       round(collected, 2),
                "lens_revenue":    round(lens_rev, 2),
                "frame_revenue":   round(frame_rev, 2),
                "checkup_revenue": round(checkup_rev, 2),
                "custom_sales":    round(custom_rev, 2),
                "restock_cost":    round(restock_cost, 2),
                "op_costs":        round(op_est, 2),
                "gross_profit":    gross_profit,
            }
        })
    except Exception as e:
        return err(e)


# ══ WEEKLY SUMMARY ════════════════════════════════════════════════════════════

@app.route("/api/weekly-summary", methods=["GET"])
def weekly_summary():
    try:
        week_ago   = (date.today() - timedelta(days=7)).isoformat()
        db         = get_db()
        week_pts   = sb_rows(db.table("patients").select("total_cost")
                               .gte("admission_date", week_ago).execute())
        week_sales = sum(_float(p.get("total_cost")) for p in week_pts)
        op_costs   = sb_rows(db.table("op_costs").select("amount,frequency").execute())
        op_est     = 0.0
        for r in op_costs:
            amt = _float(r.get("amount")); freq = _str(r.get("frequency"))
            if freq=="weekly":    op_est+=amt
            elif freq=="monthly": op_est+=amt/4
            elif freq=="yearly":  op_est+=amt/52
            elif freq=="daily":   op_est+=amt*7
            elif freq=="once":    op_est+=amt/52
        return jsonify({"sales":week_sales,"op_est":round(op_est,2),"net":round(week_sales-op_est,2)})
    except Exception as e:
        return err(e)


# ══ LEDGER ════════════════════════════════════════════════════════════════════

@app.route("/api/ledger", methods=["GET"])
def get_ledger():
    try:
        return jsonify(sb_rows(get_db().table("ledger").select("*")
                                .order("date",desc=True).order("id",desc=True).execute()))
    except Exception as e:
        return err(e)


@app.route("/api/ledger", methods=["POST"])
def save_ledger():
    try:
        d = request.json or {}
        get_db().table("ledger").insert({
            "date":         _str(d.get("date")) or today(),
            "entry_type":   _str(d.get("entry_type"),"expense"),
            "category":     _str(d.get("category")),
            "description":  _str(d.get("description")),
            "total_amount": _float(d.get("total_amount")),
            "paid_amount":  _float(d.get("paid_amount")),
            "is_expense":   _bool(d.get("is_expense"),True),
            "source_ref":   _str(d.get("source_ref")),
        }).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return err(e)


@app.route("/api/ledger/<int:lid>", methods=["DELETE"])
def delete_ledger(lid):
    try:
        get_db().table("ledger").delete().eq("id",lid).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return err(e)


# ══ CUSTOM SALES ══════════════════════════════════════════════════════════════

@app.route("/api/custom-sales", methods=["GET"])
def get_custom_sales():
    try:
        start = request.args.get("start",(date.today()-timedelta(days=30)).isoformat())
        end   = request.args.get("end",today())
        return jsonify(sb_rows(get_db().table("custom_sales").select("*")
                                .gte("date",start).lte("date",end)
                                .order("date",desc=True).order("id",desc=True).execute()))
    except Exception as e:
        return err(e)


@app.route("/api/custom-sales", methods=["POST"])
def save_custom_sale():
    try:
        d         = request.json or {}
        sid       = _int(d.get("id"))
        qty       = _int(d.get("qty"),1)
        price     = _float(d.get("unit_price"))
        total     = round(qty*price,2)
        sale_date = _str(d.get("date")) or today()
        item_name = _str(d.get("item_name"))
        category  = _str(d.get("category")) or "متنوعة"
        notes     = _str(d.get("notes"))
        lens_id   = _int(d.get("lens_id"))
        db        = get_db()
        payload   = {"date":sale_date,"item_name":item_name,"category":category,
                     "qty":qty,"unit_price":price,"total":total,"notes":notes,
                     "lens_id":lens_id or None}
        if sid:
            db.table("custom_sales").update(payload).eq("id",sid).execute()
            db.table("ledger").update({"date":sale_date,"description":f"{item_name} × {qty}",
                                       "total_amount":total,"paid_amount":total,"category":category}
                                     ).eq("source_ref",f"custom_sale:{sid}").execute()
        else:
            res = db.table("custom_sales").insert(payload).execute()
            sid = sb_one(res)["id"]
            db.table("ledger").insert({"date":sale_date,"entry_type":"income","category":category,
                "description":f"{item_name} × {qty}","total_amount":total,"paid_amount":total,
                "is_expense":False,"source_ref":f"custom_sale:{sid}"}).execute()
            if lens_id:
                try:
                    row = sb_one(db.table("lenses").select("stock").eq("id",lens_id).execute())
                    if row:
                        new_stock = max(0, _int(row.get("stock"),0) - qty)
                        db.table("lenses").update({"stock": new_stock}).eq("id",lens_id).execute()
                except Exception:
                    pass
        return jsonify({"ok":True,"id":sid})
    except Exception as e:
        return err(e)


@app.route("/api/custom-sales/<int:sid>", methods=["DELETE"])
def delete_custom_sale(sid):
    try:
        db = get_db()
        # Fetch sale before deleting so we can restore stock
        sale_rows = sb_rows(db.table("custom_sales").select("lens_id,qty").eq("id",sid).execute())
        db.table("custom_sales").delete().eq("id",sid).execute()
        db.table("ledger").delete().eq("source_ref",f"custom_sale:{sid}").execute()
        # Restore lens stock if this sale deducted from inventory
        if sale_rows:
            lid = _int(sale_rows[0].get("lens_id"))
            qty = _int(sale_rows[0].get("qty"), 1)
            if lid and qty > 0:
                try:
                    row = sb_one(db.table("lenses").select("stock").eq("id",lid).execute())
                    if row:
                        db.table("lenses").update({"stock": _int(row.get("stock"),0) + qty}).eq("id",lid).execute()
                except Exception:
                    pass
        return jsonify({"ok":True})
    except Exception as e:
        return err(e)


# ══════════════════════════════════════════════════════════════════════════════
# BACKUP / RESTORE
# GET  /api/backup   → download alnoor_backup_TIMESTAMP.zip  (full JSON dump)
# POST /api/restore  → upload that ZIP to wipe-and-restore
# ══════════════════════════════════════════════════════════════════════════════

BACKUP_TABLES = [
    "settings",
    "lenses",       "lens_restocks",
    "op_costs",
    "patients",     "patient_payments",
    "debtors",      "debtor_payments",
    "custom_sales", "ledger",
]

@app.route("/api/backup", methods=["GET"])
def backup():
    try:
        db   = get_db()
        dump = {}
        for tbl in BACKUP_TABLES:
            try:
                dump[tbl] = sb_rows(db.table(tbl).select("*").execute())
            except Exception as te:
                dump[tbl] = []
                print(f"backup skip {tbl}: {te}")

        ts         = datetime.now().strftime("%Y%m%d_%H%M%S")
        json_bytes = json.dumps(dump, ensure_ascii=False, indent=2).encode("utf-8")
        buf        = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(f"alnoor_backup_{ts}.json", json_bytes)
        buf.seek(0)
        return send_file(buf, mimetype="application/zip", as_attachment=True,
                         download_name=f"alnoor_backup_{ts}.zip")
    except Exception as e:
        return err(e)


@app.route("/api/restore", methods=["POST"])
def restore():
    try:
        f = request.files.get("file")
        if not f:
            return jsonify({"error": "No file uploaded"}), 400
        buf = io.BytesIO(f.read())
        with zipfile.ZipFile(buf) as zf:
            jname = next((n for n in zf.namelist() if n.endswith(".json")), None)
            if not jname:
                return jsonify({"error": "No JSON inside ZIP"}), 400
            dump = json.loads(zf.read(jname).decode("utf-8"))

        db      = get_db()
        summary = {}
        for tbl in BACKUP_TABLES:
            rows = dump.get(tbl, [])
            # Wipe existing rows
            try:
                if tbl == "settings":
                    db.table(tbl).delete().neq("key","").execute()
                else:
                    db.table(tbl).delete().neq("id",0).execute()
            except Exception as de:
                print(f"restore wipe {tbl}: {de}")
            if not rows:
                summary[tbl] = 0
                continue
            for i in range(0, len(rows), 100):
                db.table(tbl).insert(rows[i:i+100]).execute()
            summary[tbl] = len(rows)
        return jsonify({"ok": True, "restored": summary})
    except Exception as e:
        return err(e)



# ══ EXCEL EXPORT ══════════════════════════════════════════════════════════════

@app.route("/api/export/excel", methods=["GET"])
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        db = get_db()

        wb = openpyxl.Workbook()

        # ── helpers ──────────────────────────────────────────────────────────
        HDR_FILL  = PatternFill("solid", fgColor="1E3A5F")
        HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
        ALT_FILL  = PatternFill("solid", fgColor="F0F4FA")
        THIN      = Side(style="thin", color="D0D5DD")
        BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
        RIGHT     = Alignment(horizontal="right", vertical="center")

        def style_header(ws, headers):
            ws.append(headers)
            for col, _ in enumerate(headers, 1):
                cell = ws.cell(1, col)
                cell.fill  = HDR_FILL
                cell.font  = HDR_FONT
                cell.alignment = CENTER
                cell.border = BORDER
                ws.column_dimensions[get_column_letter(col)].width = max(14, len(str(headers[col-1]))*2)

        def style_row(ws, row_idx, num_cols):
            fill = ALT_FILL if row_idx % 2 == 0 else None
            for col in range(1, num_cols+1):
                cell = ws.cell(row_idx, col)
                if fill: cell.fill = fill
                cell.border = BORDER
                cell.alignment = RIGHT

        def fmt_date(v):
            if not v: return ""
            try: return str(v)[:10]
            except: return str(v)

        def iq(v):
            try: return int(float(v or 0))
            except: return 0

        # ── Sheet 1: Patients ──────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "سجل المراجعين"
        hdrs1 = ["التاريخ","الاسم","العمر","الهاتف","نوع الوصفة","إجمالي البيع","المدفوع","المتبقي","حالة الدفع","كشف طبي","العنوان","ملاحظات"]
        style_header(ws1, hdrs1)
        patients = sb_rows(db.table("patients").select("*").order("admission_date", desc=True).execute())
        for i, p in enumerate(patients, 2):
            total  = iq(p.get("total_cost") or p.get("total_sale"))
            paid   = iq(p.get("amount_paid"))
            rem    = total - paid
            status = "مدفوع بالكامل" if rem <= 0 else ("دفع جزئي" if paid > 0 else "لم يدفع")
            checkup = "نعم" if p.get("checkup_done") else "لا"
            row = [
                fmt_date(p.get("admission_date")), p.get("name",""), str(p.get("age","")),
                str(p.get("phone","")), str(p.get("rxtype","")),
                total, paid, rem, status, checkup,
                str(p.get("address","")), str(p.get("notes",""))
            ]
            ws1.append(row)
            style_row(ws1, i, len(hdrs1))
        ws1.freeze_panes = "A2"

        # ── Sheet 2: Lenses ────────────────────────────────────────────────
        ws2 = wb.create_sheet("مخزون العدسات")
        hdrs2 = ["نوع العدسة","الفئة","التصفية","المادة","SPH","CYL","المخزون","نقطة الإعادة","التكلفة","سعر البيع","الهامش%"]
        style_header(ws2, hdrs2)
        lenses = sb_rows(db.table("lenses").select("*").order("lens_type").execute())
        for i, l in enumerate(lenses, 2):
            cost  = _float(l.get("cost"))
            price = _float(l.get("price"))
            margin = round((price-cost)/price*100) if price else 0
            row = [
                str(l.get("lens_type","")), str(l.get("category","")),
                str(l.get("filter_type","")), str(l.get("material","")),
                str(l.get("sph","")), str(l.get("cyl","")),
                _int(l.get("stock")), _int(l.get("reorder_point"),5),
                cost, price, margin
            ]
            ws2.append(row)
            style_row(ws2, i, len(hdrs2))
        ws2.freeze_panes = "A2"

        # ── Sheet 3: Custom Sales ──────────────────────────────────────────
        ws3 = wb.create_sheet("المبيعات المتنوعة")
        hdrs3 = ["التاريخ","المنتج","الفئة","الكمية","سعر الوحدة","الإجمالي"]
        style_header(ws3, hdrs3)
        csales = sb_rows(db.table("custom_sales").select("*").order("date", desc=True).execute())
        for i, s in enumerate(csales, 2):
            qty   = _int(s.get("qty"),1)
            price = _float(s.get("price"))
            row = [
                fmt_date(s.get("date")), str(s.get("product","")),
                str(s.get("category","")), qty, price, round(qty*price)
            ]
            ws3.append(row)
            style_row(ws3, i, len(hdrs3))
        ws3.freeze_panes = "A2"

        # ── Sheet 4: Op Costs ──────────────────────────────────────────────
        ws4 = wb.create_sheet("التكاليف التشغيلية")
        hdrs4 = ["التاريخ","البند","الفئة","المبلغ","ملاحظات"]
        style_header(ws4, hdrs4)
        opcosts = sb_rows(db.table("op_costs").select("*").order("date", desc=True).execute())
        for i, o in enumerate(opcosts, 2):
            row = [
                fmt_date(o.get("date")), str(o.get("description","")),
                str(o.get("category","")), _float(o.get("amount")), str(o.get("notes",""))
            ]
            ws4.append(row)
            style_row(ws4, i, len(hdrs4))
        ws4.freeze_panes = "A2"

        # ── Save & send ────────────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        from datetime import date as _date
        fname = f"noor_clinic_{_date.today().isoformat()}.xlsx"
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=fname
        )
    except ImportError:
        return jsonify({"error": "openpyxl not installed. Add openpyxl to requirements.txt and redeploy."}), 503
    except Exception as e:
        return err(e)


# ══ AI PROXY ══════════════════════════════════════════════════════════════════

@app.route("/api/ai", methods=["POST"])
def ai_proxy():
    api_key = os.environ.get("ANTHROPIC_API_KEY","").strip()
    if not api_key:
        return jsonify({"error":"no_api_key","message":"ANTHROPIC_API_KEY not set."}), 503
    body    = request.json or {}
    payload = json.dumps({"model":body.get("model","claude-haiku-4-5-20251001"),
                          "max_tokens":body.get("max_tokens",500),
                          "messages":body.get("messages",[])}).encode()
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages", data=payload,
        headers={"Content-Type":"application/json","x-api-key":api_key,
                 "anthropic-version":"2023-06-01"}, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            return app.response_class(response=resp.read(),status=resp.status,mimetype="application/json")
    except urllib.error.HTTPError as exc:
        return app.response_class(response=exc.read(),status=exc.code,mimetype="application/json")
    except Exception as exc:
        return jsonify({"error":str(exc)}), 500


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT",5000)), debug=False)
