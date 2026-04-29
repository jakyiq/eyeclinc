#!/usr/bin/env python3
"""
app.py — Flask application — Noor Optical Clinic SaaS v3
Adds: white-label (clinic_name from DB), self-signup flow, owner notifications
"""

import os
import io
import re
import json
import secrets
import zipfile
import hashlib
import traceback
import urllib.request
import urllib.error
from datetime import datetime, date, timedelta
from functools import wraps
from collections import defaultdict

import bcrypt
from flask import Flask, request, jsonify, send_from_directory, send_file
from supabase import create_client, Client

# ── App setup ──────────────────────────────────────────────────────────────────
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR = os.path.join(BASE_DIR, "static")
os.makedirs(STATIC_DIR, exist_ok=True)

app = Flask(__name__, static_folder=STATIC_DIR)
app.config["MAX_CONTENT_LENGTH"] = 32 * 1024 * 1024

# ── Supabase client ────────────────────────────────────────────────────────────
# IMPORTANT: Use the SERVICE ROLE key here, not the anon key.
# The service role key bypasses RLS so the backend has full access.
SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "")   # must be service_role key

_supabase: Client = None

def get_db() -> Client:
    global _supabase
    if _supabase is None:
        _supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
    return _supabase

def init_db():
    pass  # tables created via schema_v2.sql in Supabase SQL editor

# ── Signup rate limiter (in-memory: max 5 per IP per hour) ────────────────────
_signup_attempts: dict = defaultdict(list)   # ip -> [timestamp, ...]

def _signup_rate_ok(ip: str) -> bool:
    now = datetime.utcnow()
    cutoff = now - timedelta(hours=1)
    _signup_attempts[ip] = [t for t in _signup_attempts[ip] if t > cutoff]
    if len(_signup_attempts[ip]) >= 5:
        return False
    _signup_attempts[ip].append(now)
    return True

# ── Notification helper (Telegram or email via Resend) ────────────────────────
def _notify_owner(clinic_name: str, username: str, phone: str, ip: str):
    """Best-effort notification to the owner when a new clinic signs up."""
    msg = (
        f"🆕 عيادة جديدة اشتركت!\n"
        f"الاسم: {clinic_name}\n"
        f"المستخدم: {username}\n"
        f"الهاتف: {phone or '—'}\n"
        f"الوقت: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC\n"
        f"IP: {ip or '—'}"
    )
    # ── Option A: Telegram ────────────────────────────────────────────────────
    tg_token   = os.environ.get("TELEGRAM_BOT_TOKEN", "").strip()
    tg_chat_id = os.environ.get("TELEGRAM_CHAT_ID", "").strip()
    if tg_token and tg_chat_id:
        try:
            payload = json.dumps({"chat_id": tg_chat_id, "text": msg}).encode()
            req = urllib.request.Request(
                f"https://api.telegram.org/bot{tg_token}/sendMessage",
                data=payload, headers={"Content-Type": "application/json"}, method="POST")
            urllib.request.urlopen(req, timeout=5)
        except Exception as e:
            print("Telegram notify failed:", e)
    # ── Option B: Resend email ────────────────────────────────────────────────
    resend_key  = os.environ.get("RESEND_API_KEY", "").strip()
    owner_email = os.environ.get("OWNER_EMAIL", "").strip()
    if resend_key and owner_email:
        try:
            body = json.dumps({
                "from":    "noreply@notifications.noor-clinic.app",
                "to":      [owner_email],
                "subject": f"عيادة جديدة: {clinic_name}",
                "text":    msg,
            }).encode()
            req = urllib.request.Request(
                "https://api.resend.com/emails", data=body,
                headers={"Content-Type": "application/json",
                         "Authorization": f"Bearer {resend_key}"}, method="POST")
            urllib.request.urlopen(req, timeout=5)
        except Exception as e:
            print("Resend notify failed:", e)

def _get_clinic_name(clinic_id: int) -> str:
    """Return the clinic's display name from the settings table, fall back to clinics.name."""
    try:
        db = get_db()
        row = db.table("settings").select("value") \
                .eq("clinic_id", clinic_id).eq("key", "clinic_name") \
                .limit(1).execute()
        if row.data:
            return row.data[0]["value"] or ""
        # fall back to clinics table
        c = db.table("clinics").select("name").eq("id", clinic_id).limit(1).execute()
        return c.data[0]["name"] if c.data else ""
    except Exception:
        return ""

# ── Type helpers ───────────────────────────────────────────────────────────────
def _str(v, default=""):
    if v is None: return default
    return str(v).strip()

def _int(v, default=None):
    try:    return int(v) if v not in (None, "", "null") else default
    except: return default

def _float(v, default=0.0):
    try:    return float(v) if v not in (None, "", "null") else default
    except: return default

def _bool(v, default=False):
    if isinstance(v, bool): return v
    if isinstance(v, int):  return bool(v)
    if isinstance(v, str):  return v.strip().lower() in ("1","true","yes")
    return default

def today():
    return date.today().isoformat()

def calc_next_review(d=None):
    try:
        dt = datetime.strptime(d or today(), "%Y-%m-%d")
        m  = dt.month + 6
        y  = dt.year + (m - 1) // 12
        m  = (m - 1) % 12 + 1
        dt = dt.replace(year=y, month=m)
    except Exception:
        dt = datetime.now() + timedelta(days=180)
    return dt.strftime("%Y-%m-%d")

def sb_rows(response):
    return response.data or []

def sb_one(response):
    data = response.data
    return data[0] if data else None

def err(e):
    tb = traceback.format_exc()
    print("ERROR:", tb)
    return jsonify({"error": str(e)}), 400


# ══════════════════════════════════════════════════════════════════════════════
# AUTH HELPERS
# ══════════════════════════════════════════════════════════════════════════════

SESSION_HOURS = 8       # session lifetime
TOKEN_HEADER  = "X-Session-Token"

def _get_token():
    """Extract session token from header or cookie."""
    return (request.headers.get(TOKEN_HEADER)
            or request.cookies.get("session_token")
            or "")

def _resolve_session(token: str):
    """Return session row or None. Also cleans expired session."""
    if not token:
        return None
    db = get_db()
    row = sb_one(
        db.table("sessions")
          .select("*")
          .eq("token", token)
          .gt("expires_at", datetime.utcnow().isoformat())
          .execute()
    )
    return row

def _require_roles(*roles):
    """Decorator factory — protects a route to specific roles."""
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            session = _resolve_session(_get_token())
            if not session:
                return jsonify({"error": "unauthenticated"}), 401
            if roles and session["role"] not in roles:
                return jsonify({"error": "forbidden",
                                "message": "ليس لديك صلاحية لهذا الإجراء"}), 403
            # Inject session into g-like request context
            request.session = session
            return fn(*args, **kwargs)
        return wrapper
    return decorator

# Convenience aliases
def _auth(*roles):
    """Shorthand: _auth() = any logged-in user; _auth('admin') = admins only."""
    return _require_roles(*roles)

def _license_ok(clinic_id: int) -> dict:
    """
    Returns {"ok": True} or {"ok": False, "message": "...", "grace": bool}
    Grace period = license expired but within grace_days.
    """
    db = get_db()
    row = sb_one(
        db.table("licenses")
          .select("*")
          .eq("clinic_id", clinic_id)
          .eq("is_active", True)
          .order("expires_at", desc=True)
          .limit(1)
          .execute()
    )
    if not row:
        return {"ok": False, "grace": False,
                "message": "لا يوجد ترخيص نشط لهذه العيادة"}

    expires = date.fromisoformat(row["expires_at"])
    grace   = int(row.get("grace_days") or 5)
    today_d = date.today()

    if today_d <= expires:
        days_left = (expires - today_d).days
        return {"ok": True, "grace": False,
                "expires": str(expires), "days_left": days_left,
                "plan": row["plan"]}

    grace_end = expires + timedelta(days=grace)
    if today_d <= grace_end:
        days_late = (today_d - expires).days
        return {"ok": True, "grace": True,
                "expires": str(expires), "days_late": days_late,
                "grace_days": grace,
                "message": f"⚠ انتهى الاشتراك منذ {days_late} يوم. يرجى التجديد خلال {grace - days_late} يوم."}

    return {"ok": False, "grace": False,
            "expires": str(expires),
            "message": "انتهت صلاحية الترخيص. يرجى تجديد الاشتراك للمتابعة."}

def _audit(action: str, detail: str = "", session=None):
    """Write an audit log entry (best-effort — never raises)."""
    try:
        db = get_db()
        db.table("audit_log").insert({
            "clinic_id": session["clinic_id"] if session else 1,
            "user_id":   session["user_id"]   if session else None,
            "username":  session.get("username","") if session else "",
            "action":    action,
            "detail":    detail,
            "ip":        request.remote_addr or "",
        }).execute()
    except Exception:
        pass


# ══════════════════════════════════════════════════════════════════════════════
# AUTH ROUTES  (no token required — these ARE the login endpoints)
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/auth/login", methods=["POST"])
def login():
    d        = request.json or {}
    username = _str(d.get("username")).lower()
    password = _str(d.get("password"))
    if not username or not password:
        return jsonify({"error": "أدخل اسم المستخدم وكلمة المرور"}), 400

    db  = get_db()
    row = sb_one(
        db.table("users")
          .select("*")
          .eq("username", username)
          .eq("is_active", True)
          .execute()
    )
    if not row:
        return jsonify({"error": "اسم المستخدم أو كلمة المرور غير صحيحة"}), 401

    # Verify bcrypt password
    try:
        pw_ok = bcrypt.checkpw(password.encode(), row["password_hash"].encode())
    except Exception:
        pw_ok = False

    if not pw_ok:
        return jsonify({"error": "اسم المستخدم أو كلمة المرور غير صحيحة"}), 401

    # Check license
    lic = _license_ok(row["clinic_id"])
    if not lic["ok"]:
        return jsonify({"error": lic["message"], "license_expired": True}), 403

    # Create session token
    token      = secrets.token_hex(32)
    expires_at = (datetime.utcnow() + timedelta(hours=SESSION_HOURS)).isoformat()

    db.table("sessions").insert({
        "token":      token,
        "user_id":    row["id"],
        "clinic_id":  row["clinic_id"],
        "role":       row["role"],
        "expires_at": expires_at,
    }).execute()

    # Update last_login
    db.table("users").update({"last_login": datetime.utcnow().isoformat()}).eq("id", row["id"]).execute()

    _audit("login", f"role={row['role']}")

    resp = jsonify({
        "ok":        True,
        "token":     token,
        "role":      row["role"],
        "full_name": row.get("full_name",""),
        "username":  row["username"],
        "clinic_id": row["clinic_id"],
        "license":   lic,
    })
    resp.set_cookie("session_token", token, httponly=True,
                    samesite="Lax", max_age=SESSION_HOURS * 3600)
    return resp


@app.route("/api/auth/logout", methods=["POST"])
def logout():
    token = _get_token()
    if token:
        try:
            get_db().table("sessions").delete().eq("token", token).execute()
        except Exception:
            pass
    resp = jsonify({"ok": True})
    resp.delete_cookie("session_token")
    return resp


@app.route("/api/auth/me", methods=["GET"])
def auth_me():
    """Returns current session info — used by frontend on page load."""
    session = _resolve_session(_get_token())
    if not session:
        return jsonify({"authenticated": False}), 200

    lic         = _license_ok(session["clinic_id"])
    clinic_name = _get_clinic_name(session["clinic_id"])
    return jsonify({
        "authenticated": True,
        "role":        session["role"],
        "clinic_id":   session["clinic_id"],
        "user_id":     session["user_id"],
        "username":    session.get("username", ""),
        "clinic_name": clinic_name,
        "license":     lic,
    })


# ══════════════════════════════════════════════════════════════════════════════
# USER MANAGEMENT  (admin only)
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/users", methods=["GET"])
@_auth("admin")
def get_users():
    db   = get_db()
    rows = sb_rows(
        db.table("users")
          .select("id,username,full_name,role,is_active,last_login,created_at")
          .eq("clinic_id", request.session["clinic_id"])
          .order("created_at")
          .execute()
    )
    return jsonify(rows)


@app.route("/api/users", methods=["POST"])
@_auth("admin")
def save_user():
    d        = request.json or {}
    uid      = _int(d.get("id"))
    username = _str(d.get("username")).lower()
    role     = _str(d.get("role"))
    db       = get_db()

    if role not in ("admin","doctor","receptionist"):
        return jsonify({"error": "دور غير صالح"}), 400

    if uid:
        # Update existing user
        payload = {
            "full_name": _str(d.get("full_name")),
            "role":      role,
            "is_active": _bool(d.get("is_active", True)),
        }
        if d.get("password"):
            pw_hash = bcrypt.hashpw(d["password"].encode(), bcrypt.gensalt(12)).decode()
            payload["password_hash"] = pw_hash
        db.table("users").update(payload).eq("id", uid).eq("clinic_id", request.session["clinic_id"]).execute()
        _audit("update_user", f"user_id={uid} role={role}", request.session)
        return jsonify({"ok": True, "id": uid})
    else:
        # Create new user
        password = _str(d.get("password"))
        if not password or len(password) < 6:
            return jsonify({"error": "كلمة المرور يجب أن تكون 6 أحرف على الأقل"}), 400
        if not username:
            return jsonify({"error": "أدخل اسم المستخدم"}), 400

        pw_hash = bcrypt.hashpw(password.encode(), bcrypt.gensalt(12)).decode()
        res = db.table("users").insert({
            "clinic_id":     request.session["clinic_id"],
            "username":      username,
            "password_hash": pw_hash,
            "full_name":     _str(d.get("full_name")),
            "role":          role,
            "is_active":     True,
        }).execute()
        new_id = sb_one(res)["id"] if sb_one(res) else None
        _audit("create_user", f"username={username} role={role}", request.session)
        return jsonify({"ok": True, "id": new_id})


@app.route("/api/users/<int:uid>", methods=["DELETE"])
@_auth("admin")
def delete_user(uid):
    # Prevent deleting yourself
    if uid == request.session["user_id"]:
        return jsonify({"error": "لا يمكنك حذف حسابك الخاص"}), 400
    get_db().table("users").delete().eq("id", uid).eq("clinic_id", request.session["clinic_id"]).execute()
    _audit("delete_user", f"user_id={uid}", request.session)
    return jsonify({"ok": True})


@app.route("/api/signup", methods=["POST"])
def signup():
    """
    Public route — no @_auth.
    Creates a new clinic + admin user + 7-day trial license in one step.
    """
    ip = request.headers.get("X-Forwarded-For", request.remote_addr or "").split(",")[0].strip()

    if not _signup_rate_ok(ip):
        return jsonify({"error": "طلبات كثيرة جداً، حاول بعد ساعة"}), 429

    d            = request.json or {}
    clinic_name  = _str(d.get("clinic_name")).strip()
    username     = _str(d.get("username")).lower().strip()
    password     = _str(d.get("password"))
    phone        = _str(d.get("phone", "")).strip()

    # ── Validate ───────────────────────────────────────────────────────────────
    if not clinic_name:
        return jsonify({"error": "أدخل اسم العيادة"}), 400
    if not username or len(username) < 3:
        return jsonify({"error": "اسم المستخدم يجب أن يكون 3 أحرف على الأقل"}), 400
    if not password or len(password) < 6:
        return jsonify({"error": "كلمة المرور يجب أن تكون 6 أحرف على الأقل"}), 400
    if not re.match(r'^[a-z0-9_]+$', username):
        return jsonify({"error": "اسم المستخدم: أحرف إنجليزية وأرقام فقط"}), 400

    db = get_db()

    # ── Check username not taken globally ──────────────────────────────────────
    existing = db.table("users").select("id").eq("username", username).execute()
    if existing.data:
        return jsonify({"error": "اسم المستخدم مستخدم بالفعل، اختر اسماً آخر"}), 409

    try:
        # ── 1. Create clinic ───────────────────────────────────────────────────
        slug = re.sub(r'[^a-z0-9]', '', clinic_name.lower().replace(' ', '_'))[:30] or "clinic"
        slug = f"{slug}_{secrets.token_hex(3)}"   # ensure uniqueness
        clinic_row = db.table("clinics").insert({
            "name":      clinic_name,
            "slug":      slug,
            "is_active": True,
        }).execute()
        clinic_id = clinic_row.data[0]["id"]

        # ── 2. Seed clinic_name into settings ─────────────────────────────────
        db.table("settings").insert({
            "clinic_id": clinic_id,
            "key":       "clinic_name",
            "value":     clinic_name,
        }).execute()

        # ── 3. Create admin user ───────────────────────────────────────────────
        pw_hash = bcrypt.hashpw(password.encode(), bcrypt.gensalt(rounds=12)).decode()
        user_row = db.table("users").insert({
            "clinic_id":     clinic_id,
            "username":      username,
            "password_hash": pw_hash,
            "full_name":     f"مدير {clinic_name}",
            "role":          "admin",
            "is_active":     True,
        }).execute()
        user_id = user_row.data[0]["id"]

        # ── 4. Create 7-day trial license ─────────────────────────────────────
        db.table("licenses").insert({
            "clinic_id":  clinic_id,
            "plan":       "trial",
            "price_iqd":  0,
            "starts_at":  date.today().isoformat(),
            "expires_at": (date.today() + timedelta(days=7)).isoformat(),
            "grace_days": 5,
            "is_active":  True,
            "notes":      f"تجربة مجانية 7 أيام — {phone or 'بدون هاتف'}",
        }).execute()

        # ── 5. Auto-login: create session ─────────────────────────────────────
        token      = secrets.token_hex(32)
        expires_at = (datetime.utcnow() + timedelta(hours=SESSION_HOURS)).isoformat()
        db.table("sessions").insert({
            "token":      token,
            "user_id":    user_id,
            "clinic_id":  clinic_id,
            "role":       "admin",
            "expires_at": expires_at,
        }).execute()

        # ── 6. Notify owner (best-effort, never fails the request) ─────────────
        try:
            _notify_owner(clinic_name, username, phone, ip)
        except Exception as ne:
            print("notify_owner failed:", ne)

        resp = jsonify({
            "ok":          True,
            "token":       token,
            "clinic_name": clinic_name,
            "clinic_id":   clinic_id,
            "role":        "admin",
            "username":    username,
            "trial_days":  7,
        })
        resp.set_cookie("session_token", token, httponly=True,
                        samesite="Lax", max_age=SESSION_HOURS * 3600)
        return resp

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": "حدث خطأ، حاول مرة أخرى"}), 500


@app.route("/api/auth/change-password", methods=["POST"])
@_auth()   # any authenticated user
def change_password():
    d        = request.json or {}
    old_pw   = _str(d.get("old_password"))
    new_pw   = _str(d.get("new_password"))
    if len(new_pw) < 6:
        return jsonify({"error": "كلمة المرور يجب أن تكون 6 أحرف على الأقل"}), 400

    db  = get_db()
    row = sb_one(db.table("users").select("password_hash").eq("id", request.session["user_id"]).execute())
    if not row or not bcrypt.checkpw(old_pw.encode(), row["password_hash"].encode()):
        return jsonify({"error": "كلمة المرور الحالية غير صحيحة"}), 401

    pw_hash = bcrypt.hashpw(new_pw.encode(), bcrypt.gensalt(12)).decode()
    db.table("users").update({"password_hash": pw_hash}).eq("id", request.session["user_id"]).execute()
    _audit("change_password", "", request.session)
    return jsonify({"ok": True})


# ══════════════════════════════════════════════════════════════════════════════
# LICENSE  (admin only)
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/license", methods=["GET"])
@_auth("admin")
def get_license():
    lic = _license_ok(request.session["clinic_id"])
    return jsonify(lic)


@app.route("/api/license", methods=["POST"])
@_auth("admin")
def add_license():
    """Admin adds a new license period (after receiving payment)."""
    d       = request.json or {}
    plan    = _str(d.get("plan","monthly"))
    db      = get_db()

    plan_days = {"monthly": 30, "bimonthly": 60, "biannual": 180, "trial": 30, "lifetime": 36500}
    plan_price = {"monthly": 30000, "bimonthly": 75000, "biannual": 210000, "trial": 0, "lifetime": 0}

    if plan not in plan_days:
        return jsonify({"error": "خطة غير صالحة"}), 400

    # Deactivate old licenses
    db.table("licenses").update({"is_active": False}).eq("clinic_id", request.session["clinic_id"]).execute()

    starts = date.fromisoformat(_str(d.get("starts_at")) or today())
    expires = starts + timedelta(days=plan_days[plan])

    db.table("licenses").insert({
        "clinic_id":   request.session["clinic_id"],
        "plan":        plan,
        "price_iqd":   _int(d.get("price_iqd"), plan_price.get(plan,0)),
        "starts_at":   str(starts),
        "expires_at":  str(expires),
        "grace_days":  5,
        "is_active":   True,
        "payment_ref": _str(d.get("payment_ref")),
        "notes":       _str(d.get("notes")),
    }).execute()

    _audit("add_license", f"plan={plan} expires={expires}", request.session)
    return jsonify({"ok": True, "expires": str(expires)})


# ══════════════════════════════════════════════════════════════════════════════
# STATIC FILE
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/")
def index():
    return send_from_directory(STATIC_DIR, "index.html")


# ══════════════════════════════════════════════════════════════════════════════
# PATIENTS  — receptionist/doctor/admin can create & view
#           — only doctor/admin can edit/delete
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/patients", methods=["GET"])
@_auth()
def get_patients():
    try:
        clinic_id = request.session["clinic_id"]
        rows = sb_rows(
            get_db().table("patients").select("*")
              .eq("clinic_id", clinic_id)
              .order("admission_date", desc=True).order("id", desc=True).execute()
        )
        return jsonify(rows)
    except Exception as e:
        return err(e)


@app.route("/api/patients", methods=["POST"])
@_auth()   # all roles can save/create
def save_patient():
    try:
        d         = request.json or {}
        pid       = _int(d.get("id"))
        db        = get_db()
        clinic_id = request.session["clinic_id"]
        role      = request.session["role"]

        # Receptionist cannot edit — only create new
        if pid and role == "receptionist":
            return jsonify({"error": "ليس لديك صلاحية تعديل سجلات المرضى"}), 403

        vals = {
            "clinic_id"    : clinic_id,
            "name"         : _str(d.get("name")),
            "age"          : _str(d.get("age")),
            "phone"        : _str(d.get("phone")),
            "address"      : _str(d.get("address")),
            "admission_date": _str(d.get("admission_date")) or today(),
            "next_review"  : calc_next_review(_str(d.get("admission_date")) or None),
            "notes"        : _str(d.get("notes")),
            "od_sph": _str(d.get("od_sph")),  "od_cyl": _str(d.get("od_cyl")),
            "od_axis": _str(d.get("od_axis")), "od_add": _str(d.get("od_add")),
            "od_va":  _str(d.get("od_va")),    "od_bcva": _str(d.get("od_bcva")),
            "os_sph": _str(d.get("os_sph")),   "os_cyl": _str(d.get("os_cyl")),
            "os_axis": _str(d.get("os_axis")), "os_add": _str(d.get("os_add")),
            "os_va":  _str(d.get("os_va")),    "os_bcva": _str(d.get("os_bcva")),
            "pd"           : _str(d.get("pd")),
            "checkup_done" : 1 if _bool(d.get("checkup_done")) else 0,
            "checkup_paid" : 1 if _bool(d.get("checkup_paid")) else 0,
            "checkup_fee"  : _float(d.get("checkup_fee"), 5000.0),
            "lenses_qty"   : _int(d.get("lenses_qty"), 2),
            "lens1_id"     : _int(d.get("lens1_id")),
            "lens1_cost"   : _float(d.get("lens1_cost")),
            "lens2_id"     : _int(d.get("lens2_id")),
            "lens2_cost"   : _float(d.get("lens2_cost")),
            "frame_price"  : _float(d.get("frame_price")),
            "frame_cost"   : _float(d.get("frame_cost")),
            "frame_name"   : _str(d.get("frame_name")),
            "frame_type"   : _str(d.get("frame_type")),
            "total_cost"   : _float(d.get("total_cost")),
            "amount_paid"  : _float(d.get("amount_paid")),
            "rxtype"       : _str(d.get("rxtype")),
            "material"     : _str(d.get("material")),
            "features"     : _str(d.get("features")),
        }

        new_l1 = vals["lens1_id"]
        new_l2 = vals["lens2_id"]
        lqty   = vals["lenses_qty"] or 2

        if pid:
            old = sb_one(db.table("patients").select("lens1_id,lens2_id,lenses_qty")
                           .eq("id", pid).eq("clinic_id", clinic_id).execute())
            old_l1 = _int(old.get("lens1_id")) if old else None
            old_l2 = _int(old.get("lens2_id")) if old else None
            db.table("patients").update(vals).eq("id", pid).eq("clinic_id", clinic_id).execute()
            if old_l1 and old_l1 != new_l1: _lens_adjust(db, old_l1, +1, clinic_id)
            if old_l2 and old_l2 != new_l2: _lens_adjust(db, old_l2, +1, clinic_id)
            if new_l1 and new_l1 != old_l1: _lens_adjust(db, new_l1, -1, clinic_id)
            if new_l2 and lqty == 2 and new_l2 != old_l2: _lens_adjust(db, new_l2, -1, clinic_id)
            # Sync ledger
            try:
                db.table("ledger").update({
                    "date": vals["admission_date"],
                    "description": f"بيع للمريض: {vals['name']}",
                    "total_amount": vals["total_cost"],
                    "paid_amount": vals["amount_paid"],
                }).eq("source_ref", f"patient:{pid}").eq("is_expense", False).execute()
            except Exception:
                pass
        else:
            res = db.table("patients").insert(vals).execute()
            row = sb_one(res)
            if not row:
                return jsonify({"error": "Insert failed — check Supabase RLS"}), 400
            pid = row["id"]
            if new_l1: _lens_adjust(db, new_l1, -1, clinic_id)
            if new_l2 and lqty == 2: _lens_adjust(db, new_l2, -1, clinic_id)
            if vals["total_cost"] > 0:
                db.table("ledger").insert({
                    "clinic_id": clinic_id,
                    "date": vals["admission_date"], "entry_type": "income",
                    "category": "مبيعات مرضى",
                    "description": f"بيع للمريض: {vals['name']}",
                    "total_amount": vals["total_cost"], "paid_amount": vals["amount_paid"],
                    "is_expense": False, "source_ref": f"patient:{pid}",
                }).execute()

        return jsonify({"id": pid})
    except Exception as e:
        return err(e)


def _lens_adjust(db, lid, delta, clinic_id=1):
    try:
        row = sb_one(db.table("lenses").select("stock").eq("id", lid).eq("clinic_id", clinic_id).execute())
        if row:
            db.table("lenses").update(
                {"stock": max(0, (_int(row.get("stock"), 0) or 0) + delta)}
            ).eq("id", lid).execute()
    except Exception as e:
        print(f"_lens_adjust: {e}")


@app.route("/api/patients/<int:pid>", methods=["DELETE"])
@_auth("admin", "doctor")
def delete_patient(pid):
    try:
        db        = get_db()
        clinic_id = request.session["clinic_id"]
        row = sb_one(db.table("patients").select("*").eq("id", pid).eq("clinic_id", clinic_id).execute())
        if row:
            if row.get("lens1_id"): _lens_adjust(db, row["lens1_id"], +1, clinic_id)
            if row.get("lens2_id") and (_int(row.get("lenses_qty"), 2) or 2) == 2:
                _lens_adjust(db, row["lens2_id"], +1, clinic_id)
            try:
                db.table("ledger").delete().eq("source_ref", f"patient:{pid}").execute()
            except Exception:
                pass
        db.table("patients").delete().eq("id", pid).eq("clinic_id", clinic_id).execute()
        _audit("delete_patient", f"patient_id={pid} name={row.get('name','') if row else ''}", request.session)
        return jsonify({"ok": True})
    except Exception as e:
        return err(e)


@app.route("/api/patients/<int:pid>/pay", methods=["POST"])
@_auth()
def pay_patient(pid):
    try:
        d         = request.json or {}
        amount    = _float(d.get("amount"))
        db        = get_db()
        clinic_id = request.session["clinic_id"]
        row = sb_one(db.table("patients").select("amount_paid").eq("id", pid).eq("clinic_id", clinic_id).execute())
        if row:
            db.table("patients").update(
                {"amount_paid": _float(row.get("amount_paid")) + amount}
            ).eq("id", pid).execute()
        db.table("patient_payments").insert({
            "clinic_id": clinic_id, "patient_id": pid, "amount": amount,
            "date": _str(d.get("date")) or today(),
        }).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return err(e)


# ══════════════════════════════════════════════════════════════════════════════
# DEBTORS  — admin only
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/debtors", methods=["GET"])
@_auth("admin")
def get_debtors():
    try:
        clinic_id = request.session["clinic_id"]
        return jsonify(sb_rows(get_db().table("debtors").select("*")
                                .eq("clinic_id", clinic_id)
                                .order("created_at", desc=True).execute()))
    except Exception as e:
        return err(e)


@app.route("/api/debtors", methods=["POST"])
@_auth("admin")
def save_debtor():
    try:
        d         = request.json or {}
        did       = _int(d.get("id"))
        total     = _float(d.get("total_we_owe"))
        init_paid = _float(d.get("initial_paid"))
        remaining = max(0.0, total - init_paid)
        status    = "settled" if remaining <= 0 else ("partial" if init_paid > 0 else "unpaid")
        db        = get_db()
        clinic_id = request.session["clinic_id"]
        payload   = {
            "clinic_id": clinic_id,
            "name": _str(d.get("name")), "phone": _str(d.get("phone")),
            "category": _str(d.get("category")), "what_bought": _str(d.get("what_bought")),
            "total_we_owe": total, "remaining": remaining, "status": status,
            "due_date": _str(d.get("due_date")) or None, "notes": _str(d.get("notes")),
        }
        if did:
            db.table("debtors").update(payload).eq("id", did).eq("clinic_id", clinic_id).execute()
        else:
            payload["total_paid"] = init_paid
            res = db.table("debtors").insert(payload).execute()
            did = sb_one(res)["id"]
        return jsonify({"id": did})
    except Exception as e:
        return err(e)


@app.route("/api/debtors/<int:did>", methods=["DELETE"])
@_auth("admin")
def delete_debtor(did):
    try:
        clinic_id = request.session["clinic_id"]
        get_db().table("debtors").delete().eq("id", did).eq("clinic_id", clinic_id).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return err(e)


@app.route("/api/debtors/<int:did>/pay", methods=["POST"])
@_auth("admin")
def pay_debtor(did):
    try:
        d         = request.json or {}
        amount    = _float(d.get("amount"))
        db        = get_db()
        clinic_id = request.session["clinic_id"]
        row = sb_one(db.table("debtors").select("total_paid,remaining").eq("id", did).eq("clinic_id", clinic_id).execute())
        if row:
            new_paid      = _float(row.get("total_paid")) + amount
            new_remaining = max(0.0, _float(row.get("remaining")) - amount)
            db.table("debtors").update({
                "total_paid": new_paid, "remaining": new_remaining,
                "status": "settled" if new_remaining <= 0 else ("partial" if new_paid > 0 else "unpaid"),
            }).eq("id", did).execute()
        db.table("debtor_payments").insert({
            "clinic_id": clinic_id, "debtor_id": did, "amount": amount,
            "method": _str(d.get("method"), "cash"),
            "note": _str(d.get("note")),
            "date": _str(d.get("date")) or today(),
        }).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return err(e)


# ══════════════════════════════════════════════════════════════════════════════
# LENSES  — doctor/admin can manage; receptionist read-only
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/lenses", methods=["GET"])
@_auth()
def get_lenses():
    try:
        clinic_id = request.session["clinic_id"]
        return jsonify(sb_rows(get_db().table("lenses").select("*")
                                .eq("clinic_id", clinic_id)
                                .order("lens_type").order("sph").execute()))
    except Exception as e:
        return err(e)


@app.route("/api/lenses", methods=["POST"])
@_auth("admin", "doctor")
def save_lens():
    try:
        d         = request.json or {}
        lid       = _int(d.get("id"))
        db        = get_db()
        clinic_id = request.session["clinic_id"]
        payload   = {
            "clinic_id": clinic_id,
            "lens_type": _str(d.get("lens_type")), "category": _str(d.get("category")),
            "sph": _str(d.get("sph")), "cyl": _str(d.get("cyl")),
            "material": _str(d.get("material")), "filter_type": _str(d.get("filter_type")),
            "tint": _str(d.get("tint")),
            "stock": _int(d.get("stock"), 0), "reorder_point": _int(d.get("reorder_point"), 5),
            "cost": _float(d.get("cost")), "price": _float(d.get("price")),
        }
        if lid: db.table("lenses").update(payload).eq("id", lid).execute()
        else:   db.table("lenses").insert(payload).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return err(e)


@app.route("/api/lenses/<int:lid>/restock", methods=["POST"])
@_auth("admin", "doctor")
def restock_lens(lid):
    try:
        d         = request.json or {}
        qty       = _int(d.get("qty"), 0)
        cost      = _float(d.get("cost"))
        db        = get_db()
        clinic_id = request.session["clinic_id"]
        row = sb_one(db.table("lenses").select("*").eq("id", lid).eq("clinic_id", clinic_id).execute())
        if not row: return jsonify({"error": "Lens not found"}), 404
        new_stock = (_int(row.get("stock"), 0) or 0) + qty
        upd = {"stock": new_stock}
        if cost > 0: upd["cost"] = cost
        db.table("lenses").update(upd).eq("id", lid).execute()
        db.table("lens_restocks").insert({
            "clinic_id": clinic_id, "lens_id": lid, "qty": qty, "cost_per_unit": cost,
            "date": _str(d.get("date")) or today(),
        }).execute()
        if qty * cost > 0:
            db.table("ledger").insert({
                "clinic_id": clinic_id,
                "date": _str(d.get("date")) or today(), "entry_type": "expense",
                "category": "شراء عدسات",
                "description": f"تجديد: {row.get('lens_type','')} SPH{row.get('sph','')} × {qty}",
                "total_amount": qty*cost, "paid_amount": qty*cost,
                "is_expense": True, "source_ref": f"restock:lens:{lid}",
            }).execute()
        return jsonify({"ok": True, "new_stock": new_stock})
    except Exception as e:
        return err(e)


@app.route("/api/lenses/<int:lid>", methods=["DELETE"])
@_auth("admin")
def delete_lens(lid):
    try:
        clinic_id = request.session["clinic_id"]
        get_db().table("lenses").delete().eq("id", lid).eq("clinic_id", clinic_id).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return err(e)


# ══════════════════════════════════════════════════════════════════════════════
# OP-COSTS  — admin only
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/op-costs", methods=["GET"])
@_auth("admin")
def get_op_costs():
    try:
        clinic_id = request.session["clinic_id"]
        return jsonify(sb_rows(get_db().table("op_costs").select("*")
                                .eq("clinic_id", clinic_id)
                                .order("created_at", desc=True).execute()))
    except Exception as e:
        return err(e)


@app.route("/api/op-costs", methods=["POST"])
@_auth("admin")
def save_op_cost():
    try:
        d         = request.json or {}
        oid       = _int(d.get("id"))
        db        = get_db()
        clinic_id = request.session["clinic_id"]
        payload   = {
            "clinic_id": clinic_id,
            "name": _str(d.get("name")), "amount": _float(d.get("amount")),
            "frequency": _str(d.get("frequency")), "category": _str(d.get("category")),
        }
        if oid: db.table("op_costs").update(payload).eq("id", oid).execute()
        else:   db.table("op_costs").insert(payload).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return err(e)


@app.route("/api/op-costs/<int:oid>", methods=["DELETE"])
@_auth("admin")
def delete_op_cost(oid):
    try:
        clinic_id = request.session["clinic_id"]
        get_db().table("op_costs").delete().eq("id", oid).eq("clinic_id", clinic_id).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return err(e)


# ══════════════════════════════════════════════════════════════════════════════
# SETTINGS  — admin only
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/settings", methods=["GET"])
@_auth()
def get_settings():
    try:
        clinic_id = request.session["clinic_id"]
        rows = sb_rows(get_db().table("settings").select("key,value")
                                .eq("clinic_id", clinic_id).execute())
        return jsonify({r["key"]: r["value"] for r in rows})
    except Exception as e:
        return err(e)


@app.route("/api/settings", methods=["POST"])
@_auth("admin")
def save_settings():
    try:
        d         = request.json or {}
        db        = get_db()
        clinic_id = request.session["clinic_id"]
        for k, v in d.items():
            db.table("settings").upsert(
                {"clinic_id": clinic_id, "key": k, "value": str(v) if v is not None else ""},
                on_conflict="clinic_id,key"
            ).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return err(e)


# ══════════════════════════════════════════════════════════════════════════════
# DASHBOARD  — all authenticated
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/dashboard", methods=["GET"])
@_auth()
def dashboard():
    try:
        db        = get_db()
        clinic_id = request.session["clinic_id"]
        role      = request.session["role"]
        t         = today()
        week_ago  = (date.today() - timedelta(days=7)).isoformat()

        today_pts = sb_rows(db.table("patients").select("total_cost,amount_paid")
                              .eq("clinic_id", clinic_id).eq("admission_date", t).execute())
        week_pts  = sb_rows(db.table("patients").select("total_cost")
                              .eq("clinic_id", clinic_id).gte("admission_date", week_ago).execute())
        all_pts   = sb_rows(db.table("patients").select("total_cost,amount_paid")
                              .eq("clinic_id", clinic_id).execute())

        op_costs  = sb_rows(db.table("op_costs").select("amount")
                              .eq("clinic_id", clinic_id).execute()) if role in ("admin","doctor") else []
        we_owe    = sb_rows(db.table("debtors").select("remaining").gt("remaining", 0)
                              .eq("clinic_id", clinic_id).execute()) if role == "admin" else []
        low_rows  = sb_rows(db.table("lenses").select("stock,reorder_point")
                              .eq("clinic_id", clinic_id).execute()) if role in ("admin","doctor") else []

        week_sales = sum(_float(p.get("total_cost")) for p in week_pts)
        total_oc   = sum(_float(r.get("amount")) for r in op_costs)

        result = {
            "today": {
                "sales":     sum(_float(p.get("total_cost"))  for p in today_pts),
                "collected": sum(_float(p.get("amount_paid")) for p in today_pts),
                "patients":  len(today_pts),
            },
            "week":      {"sales": week_sales, "net": week_sales - total_oc},
            "pt_debt":   sum(max(0.0, _float(p.get("total_cost")) - _float(p.get("amount_paid"))) for p in all_pts),
            "we_owe":    sum(_float(r.get("remaining")) for r in we_owe),
            "low_stock": sum(1 for r in low_rows if (_int(r.get("stock"),0) or 0) <= (_int(r.get("reorder_point"),5) or 5)),
        }
        # Attach license info for banner
        result["license"] = _license_ok(clinic_id)
        return jsonify(result)
    except Exception as e:
        return err(e)


# ══════════════════════════════════════════════════════════════════════════════
# REPORTS  — admin and doctor
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/reports", methods=["GET"])
@_auth("admin", "doctor")
def reports():
    try:
        start     = request.args.get("start", (date.today() - timedelta(days=7)).isoformat())
        end       = request.args.get("end", today())
        db        = get_db()
        clinic_id = request.session["clinic_id"]

        pts = sb_rows(db.table("patients").select("*")
                        .eq("clinic_id", clinic_id)
                        .gte("admission_date", start).lte("admission_date", end)
                        .order("admission_date", desc=True).execute())
        restocks = sb_rows(db.table("lens_restocks").select("*")
                             .eq("clinic_id", clinic_id)
                             .gte("date", start).lte("date", end)
                             .order("date", desc=True).execute())
        custom_sales = sb_rows(db.table("custom_sales").select("*")
                                 .eq("clinic_id", clinic_id)
                                 .gte("date", start).lte("date", end)
                                 .order("date", desc=True).execute())
        op_costs_rows = sb_rows(db.table("op_costs").select("*")
                                  .eq("clinic_id", clinic_id).execute())

        try:
            days = max(1, (datetime.strptime(end,"%Y-%m-%d") - datetime.strptime(start,"%Y-%m-%d")).days + 1)
        except Exception:
            days = 7

        op_est = 0.0
        for r in op_costs_rows:
            amt  = _float(r.get("amount"))
            freq = _str(r.get("frequency"))
            if freq=="weekly":    op_est += amt*(days/7)
            elif freq=="monthly": op_est += amt*(days/30)
            elif freq=="yearly":  op_est += amt*(days/365)
            elif freq=="daily":   op_est += amt*days
            elif freq=="once":    op_est += amt/52

        lens_rev    = sum((_float(p.get("lens1_cost")) + _float(p.get("lens2_cost"))) for p in pts)
        frame_rev   = sum(_float(p.get("frame_price") or p.get("frame_cost")) for p in pts)
        checkup_rev = sum(_float(p.get("checkup_fee") or 0) for p in pts if p.get("checkup_done"))
        custom_rev  = sum(_float(s.get("total")) for s in custom_sales)
        gross_rev   = lens_rev + frame_rev + checkup_rev + custom_rev
        collected   = sum(_float(p.get("amount_paid")) for p in pts)
        restock_cost= sum((_float(r.get("qty",0)) * _float(r.get("cost_per_unit",0))) for r in restocks)

        return jsonify({
            "patients": pts, "restocks": restocks,
            "custom_sales": custom_sales, "op_costs": op_costs_rows,
            "stats": {
                "gross_revenue":   round(gross_rev, 2),
                "collected":       round(collected, 2),
                "lens_revenue":    round(lens_rev, 2),
                "frame_revenue":   round(frame_rev, 2),
                "checkup_revenue": round(checkup_rev, 2),
                "custom_sales":    round(custom_rev, 2),
                "restock_cost":    round(restock_cost, 2),
                "op_costs":        round(op_est, 2),
                "gross_profit":    round(gross_rev - restock_cost - op_est, 2),
            }
        })
    except Exception as e:
        return err(e)


# ══════════════════════════════════════════════════════════════════════════════
# LEDGER  — admin only
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/ledger", methods=["GET"])
@_auth("admin")
def get_ledger():
    try:
        clinic_id = request.session["clinic_id"]
        return jsonify(sb_rows(get_db().table("ledger").select("*")
                                .eq("clinic_id", clinic_id)
                                .order("date",desc=True).order("id",desc=True).execute()))
    except Exception as e:
        return err(e)


@app.route("/api/ledger", methods=["POST"])
@_auth("admin")
def save_ledger():
    try:
        d         = request.json or {}
        clinic_id = request.session["clinic_id"]
        get_db().table("ledger").insert({
            "clinic_id":    clinic_id,
            "date":         _str(d.get("date")) or today(),
            "entry_type":   _str(d.get("entry_type"),"expense"),
            "category":     _str(d.get("category")),
            "description":  _str(d.get("description")),
            "total_amount": _float(d.get("total_amount")),
            "paid_amount":  _float(d.get("paid_amount")),
            "is_expense":   _bool(d.get("is_expense"),True),
            "source_ref":   _str(d.get("source_ref")),
        }).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return err(e)


@app.route("/api/ledger/<int:lid>", methods=["DELETE"])
@_auth("admin")
def delete_ledger(lid):
    try:
        clinic_id = request.session["clinic_id"]
        get_db().table("ledger").delete().eq("id",lid).eq("clinic_id",clinic_id).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return err(e)


# ══════════════════════════════════════════════════════════════════════════════
# CUSTOM SALES  — admin only
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/custom-sales", methods=["GET"])
@_auth("admin", "doctor")
def get_custom_sales():
    try:
        start     = request.args.get("start",(date.today()-timedelta(days=30)).isoformat())
        end       = request.args.get("end",today())
        clinic_id = request.session["clinic_id"]
        return jsonify(sb_rows(get_db().table("custom_sales").select("*")
                                .eq("clinic_id", clinic_id)
                                .gte("date",start).lte("date",end)
                                .order("date",desc=True).order("id",desc=True).execute()))
    except Exception as e:
        return err(e)


@app.route("/api/custom-sales", methods=["POST"])
@_auth("admin", "doctor")
def save_custom_sale():
    try:
        d         = request.json or {}
        sid       = _int(d.get("id"))
        qty       = _int(d.get("qty"),1)
        price     = _float(d.get("unit_price"))
        total     = round(qty*price,2)
        sale_date = _str(d.get("date")) or today()
        item_name = _str(d.get("item_name"))
        category  = _str(d.get("category")) or "متنوعة"
        notes     = _str(d.get("notes"))
        lens_id   = _int(d.get("lens_id"))
        db        = get_db()
        clinic_id = request.session["clinic_id"]
        payload   = {"clinic_id": clinic_id,
                     "date":sale_date,"item_name":item_name,"category":category,
                     "qty":qty,"unit_price":price,"total":total,"notes":notes,
                     "lens_id":lens_id or None}
        if sid:
            db.table("custom_sales").update(payload).eq("id",sid).execute()
            db.table("ledger").update({"date":sale_date,"description":f"{item_name} × {qty}",
                                       "total_amount":total,"paid_amount":total,"category":category}
                                     ).eq("source_ref",f"custom_sale:{sid}").execute()
        else:
            res = db.table("custom_sales").insert(payload).execute()
            sid = sb_one(res)["id"]
            db.table("ledger").insert({"clinic_id":clinic_id,"date":sale_date,"entry_type":"income","category":category,
                "description":f"{item_name} × {qty}","total_amount":total,"paid_amount":total,
                "is_expense":False,"source_ref":f"custom_sale:{sid}"}).execute()
            if lens_id:
                try:
                    row = sb_one(db.table("lenses").select("stock").eq("id",lens_id).execute())
                    if row:
                        db.table("lenses").update({"stock": max(0,_int(row.get("stock"),0)-qty)}).eq("id",lens_id).execute()
                except Exception:
                    pass
        return jsonify({"ok":True,"id":sid})
    except Exception as e:
        return err(e)


@app.route("/api/custom-sales/<int:sid>", methods=["DELETE"])
@_auth("admin")
def delete_custom_sale(sid):
    try:
        db        = get_db()
        clinic_id = request.session["clinic_id"]
        sale_rows = sb_rows(db.table("custom_sales").select("lens_id,qty")
                              .eq("id",sid).eq("clinic_id",clinic_id).execute())
        db.table("custom_sales").delete().eq("id",sid).execute()
        db.table("ledger").delete().eq("source_ref",f"custom_sale:{sid}").execute()
        if sale_rows:
            lid = _int(sale_rows[0].get("lens_id"))
            qty = _int(sale_rows[0].get("qty"), 1)
            if lid and qty > 0:
                try:
                    row = sb_one(db.table("lenses").select("stock").eq("id",lid).execute())
                    if row:
                        db.table("lenses").update({"stock": _int(row.get("stock"),0)+qty}).eq("id",lid).execute()
                except Exception:
                    pass
        return jsonify({"ok":True})
    except Exception as e:
        return err(e)


# ══════════════════════════════════════════════════════════════════════════════
# WEEKLY SUMMARY
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/weekly-summary", methods=["GET"])
@_auth("admin","doctor")
def weekly_summary():
    try:
        week_ago  = (date.today() - timedelta(days=7)).isoformat()
        db        = get_db()
        clinic_id = request.session["clinic_id"]
        week_pts  = sb_rows(db.table("patients").select("total_cost")
                              .eq("clinic_id",clinic_id).gte("admission_date", week_ago).execute())
        week_sales = sum(_float(p.get("total_cost")) for p in week_pts)
        op_costs   = sb_rows(db.table("op_costs").select("amount,frequency")
                               .eq("clinic_id",clinic_id).execute())
        op_est = 0.0
        for r in op_costs:
            amt = _float(r.get("amount")); freq = _str(r.get("frequency"))
            if freq=="weekly":    op_est+=amt
            elif freq=="monthly": op_est+=amt/4
            elif freq=="yearly":  op_est+=amt/52
            elif freq=="daily":   op_est+=amt*7
            elif freq=="once":    op_est+=amt/52
        return jsonify({"sales":week_sales,"op_est":round(op_est,2),"net":round(week_sales-op_est,2)})
    except Exception as e:
        return err(e)


# ══════════════════════════════════════════════════════════════════════════════
# BACKUP / RESTORE  — admin only
# ══════════════════════════════════════════════════════════════════════════════

BACKUP_TABLES = [
    "settings","lenses","lens_restocks","op_costs",
    "patients","patient_payments","debtors","debtor_payments",
    "custom_sales","ledger",
]

@app.route("/api/backup", methods=["GET"])
@_auth("admin")
def backup():
    try:
        db        = get_db()
        clinic_id = request.session["clinic_id"]
        dump      = {}
        for tbl in BACKUP_TABLES:
            try:
                dump[tbl] = sb_rows(db.table(tbl).select("*").eq("clinic_id",clinic_id).execute())
            except Exception as te:
                dump[tbl] = []
                print(f"backup skip {tbl}: {te}")

        ts         = datetime.now().strftime("%Y%m%d_%H%M%S")
        cn         = re.sub(r'[^\w]', '_', _get_clinic_name(clinic_id) or "clinic")[:20]
        json_bytes = json.dumps(dump, ensure_ascii=False, indent=2).encode("utf-8")
        buf        = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(f"{cn}_backup_{ts}.json", json_bytes)
        buf.seek(0)
        _audit("backup", f"tables={','.join(BACKUP_TABLES)}", request.session)
        return send_file(buf, mimetype="application/zip", as_attachment=True,
                         download_name=f"{cn}_backup_{ts}.zip")
    except Exception as e:
        return err(e)


@app.route("/api/restore", methods=["POST"])
@_auth("admin")
def restore():
    try:
        f = request.files.get("file")
        if not f: return jsonify({"error": "No file uploaded"}), 400
        buf = io.BytesIO(f.read())
        with zipfile.ZipFile(buf) as zf:
            jname = next((n for n in zf.namelist() if n.endswith(".json")), None)
            if not jname: return jsonify({"error": "No JSON inside ZIP"}), 400
            dump = json.loads(zf.read(jname).decode("utf-8"))

        db        = get_db()
        clinic_id = request.session["clinic_id"]
        summary   = {}
        for tbl in BACKUP_TABLES:
            rows = dump.get(tbl, [])
            try:
                if tbl == "settings":
                    db.table(tbl).delete().eq("clinic_id",clinic_id).execute()
                else:
                    db.table(tbl).delete().eq("clinic_id",clinic_id).execute()
            except Exception as de:
                print(f"restore wipe {tbl}: {de}")
            if not rows:
                summary[tbl] = 0
                continue
            # Ensure clinic_id is set on all rows
            for row in rows:
                row["clinic_id"] = clinic_id
            for i in range(0, len(rows), 100):
                db.table(tbl).insert(rows[i:i+100]).execute()
            summary[tbl] = len(rows)
        _audit("restore", json.dumps(summary), request.session)
        return jsonify({"ok": True, "restored": summary})
    except Exception as e:
        return err(e)


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT  — admin/doctor
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/export/excel", methods=["GET"])
@_auth("admin","doctor")
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        db        = get_db()
        clinic_id = request.session["clinic_id"]
        wb = openpyxl.Workbook()

        HDR_FILL = PatternFill("solid", fgColor="1E3A5F")
        HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
        ALT_FILL = PatternFill("solid", fgColor="F0F4FA")
        THIN     = Side(style="thin", color="D0D5DD")
        BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        CENTER   = Alignment(horizontal="center", vertical="center", wrap_text=True)
        RIGHT    = Alignment(horizontal="right",  vertical="center")

        def style_header(ws, headers):
            ws.append(headers)
            for col, _ in enumerate(headers, 1):
                cell = ws.cell(1, col)
                cell.fill  = HDR_FILL; cell.font = HDR_FONT
                cell.alignment = CENTER; cell.border = BORDER
                ws.column_dimensions[get_column_letter(col)].width = max(14, len(str(headers[col-1]))*2)

        def style_row(ws, row_idx, num_cols):
            fill = ALT_FILL if row_idx % 2 == 0 else None
            for col in range(1, num_cols+1):
                cell = ws.cell(row_idx, col)
                if fill: cell.fill = fill
                cell.border = BORDER; cell.alignment = RIGHT

        def fmt_date(v):
            if not v: return ""
            return str(v)[:10]

        ws1 = wb.active; ws1.title = "سجل المراجعين"
        hdrs1 = ["التاريخ","الاسم","العمر","الهاتف","نوع الوصفة","إجمالي البيع","المدفوع","المتبقي","حالة الدفع"]
        style_header(ws1, hdrs1)
        patients = sb_rows(db.table("patients").select("*").eq("clinic_id",clinic_id).order("admission_date",desc=True).execute())
        for i, p in enumerate(patients, 2):
            total = int(_float(p.get("total_cost"))); paid = int(_float(p.get("amount_paid"))); rem = total-paid
            status = "مدفوع" if rem<=0 else ("جزئي" if paid>0 else "غير مدفوع")
            ws1.append([fmt_date(p.get("admission_date")),p.get("name",""),str(p.get("age","")),str(p.get("phone","")),str(p.get("rxtype","")),total,paid,rem,status])
            style_row(ws1, i, len(hdrs1))
        ws1.freeze_panes = "A2"

        ws2 = wb.create_sheet("مخزون العدسات")
        hdrs2 = ["نوع العدسة","التصفية","SPH","CYL","المخزون","التكلفة","سعر البيع"]
        style_header(ws2, hdrs2)
        lenses = sb_rows(db.table("lenses").select("*").eq("clinic_id",clinic_id).order("lens_type").execute())
        for i, l in enumerate(lenses, 2):
            ws2.append([str(l.get("lens_type","")),str(l.get("filter_type","")),str(l.get("sph","")),str(l.get("cyl","")),_int(l.get("stock")),_float(l.get("cost")),_float(l.get("price"))])
            style_row(ws2, i, len(hdrs2))

        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        from datetime import date as _date
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f"noor_clinic_{_date.today().isoformat()}.xlsx")
    except ImportError:
        return jsonify({"error": "openpyxl not installed"}), 503
    except Exception as e:
        return err(e)


# ══════════════════════════════════════════════════════════════════════════════
# AI PROXY  — all authenticated users
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/ai", methods=["POST"])
@_auth()
def ai_proxy():
    api_key = os.environ.get("ANTHROPIC_API_KEY","").strip()
    if not api_key:
        return jsonify({"error":"no_api_key","message":"ANTHROPIC_API_KEY not set."}), 503
    body    = request.json or {}
    payload = json.dumps({"model":body.get("model","claude-haiku-4-5-20251001"),
                          "max_tokens":body.get("max_tokens",500),
                          "messages":body.get("messages",[])}).encode()
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages", data=payload,
        headers={"Content-Type":"application/json","x-api-key":api_key,
                 "anthropic-version":"2023-06-01"}, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            return app.response_class(response=resp.read(),status=resp.status,mimetype="application/json")
    except urllib.error.HTTPError as exc:
        return app.response_class(response=exc.read(),status=exc.code,mimetype="application/json")
    except Exception as exc:
        return jsonify({"error":str(exc)}), 500


# ── Entry point ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT",5000)), debug=False)
